# -*- coding: utf-8 -*-
"""Cruza el top10 (del reporte) con los datos extraidos de las fichas
(response time, on-time delivery, tipo de proveedor) -> XLSX enriquecido."""
import csv, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Datos extraidos de las 10 fichas (por URL id)
FICHAS = {
 1601740234896: {"resp":"≤2h","ontime":"≥93%","sup":"Multispecialty","cust":"Si"},
 1601799533956: {"resp":"≤1h","ontime":"≥99%","sup":"Multispecialty","cust":"Si"},
 1601716075635: {"resp":"≤3h","ontime":"≥99%","sup":"","cust":"Si"},
 1601295293535: {"resp":"≤4h","ontime":"≥93%","sup":"Multispecialty","cust":"Si"},
 1601424414588: {"resp":"≤2h","ontime":"≥99%","sup":"Multispecialty","cust":"Si"},
 1601023310453: {"resp":"≤3h","ontime":"≥98%","sup":"","cust":"Si"},
 1601758712138: {"resp":"≤4h","ontime":"≥90%","sup":"","cust":"Si"},
 1601765195031: {"resp":"≤3h","ontime":"≥91%","sup":"","cust":"Si"},
 1601718581780: {"resp":"≤3h","ontime":"≥91%","sup":"","cust":"Si"},
 1600238919442: {"resp":"≤2h","ontime":"≥96%","sup":"","cust":"Si"},
}

def n(s):
    m=re.search(r'\d+',str(s)); return int(m.group()) if m else 0

rows=list(csv.DictReader(open('reporte_sourcing_alibaba.csv',encoding='utf-8-sig')))
rows.sort(key=lambda r:-n(r['reviews']))
top=rows[:10]

wb=Workbook(); ws=wb.active; ws.title="Top10 enriquecido"
cols=[("#",4),("Proveedor",34),("Producto",30),("Reviews",8),("Años",6),
      ("Precio USD",11),("MOQ",14),("Response",9),("On-time",9),
      ("Tipo",14),("Custom/OEM",11),("URL",60)]
hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(color="FFFFFF",bold=True)
for i,(nm,w) in enumerate(cols,1):
    c=ws.cell(1,i,nm); c.fill=hf; c.font=hfont
    c.alignment=Alignment(horizontal="center"); ws.column_dimensions[get_column_letter(i)].width=w

for idx,r in enumerate(top,1):
    pid=int(re.search(r'_(\d+)\.html',r['url']).group(1))
    f=FICHAS.get(pid,{})
    ws.append([idx, r['company'], r['title'][:40], n(r['reviews']), n(r['years']),
               r['price'], r['moq'].replace('Min. order: ',''),
               f.get('resp',''), f.get('ontime',''), f.get('sup',''), f.get('cust',''), r['url']])
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}11"
wb.save("top10_enriquecido.xlsx")
print("OK -> top10_enriquecido.xlsx")
print(f'\n{"#":<3}{"RESP":<6}{"ONTIME":<8}{"MOQ":<14}EMPRESA')
for idx,r in enumerate(top,1):
    pid=int(re.search(r'_(\d+)\.html',r['url']).group(1)); f=FICHAS.get(pid,{})
    print(f'{idx:<3}{f.get("resp",""):<6}{f.get("ontime",""):<8}{r["moq"].replace("Min. order: ",""):<14}{r["company"][:30]}')
