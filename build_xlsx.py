# -*- coding: utf-8 -*-
"""Genera proveedores_squishy.xlsx desde reporte_sourcing_alibaba.csv, con
JOIN opcional de datos de ficha (response time, on-time, customization).
Orden: # reviews desc, anios verified desc, service desc, precio minimo asc.
"""
import csv, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\Usuario\Documents\alibaba Scrapper\reporte_sourcing_alibaba.csv"
OUT = r"C:\Users\Usuario\Documents\alibaba Scrapper\proveedores_squishy.xlsx"

# Datos de ficha (por productId). Se unen por URL; el resto quedan en blanco.
FICHAS = {
 1601740234896: {"resp": "<=2h", "ontime": ">=93%", "cust": "Si"},
 1601799533956: {"resp": "<=1h", "ontime": ">=99%", "cust": "Si"},
 1601716075635: {"resp": "<=3h", "ontime": ">=99%", "cust": "Si"},
 1601295293535: {"resp": "<=4h", "ontime": ">=93%", "cust": "Si"},
 1601424414588: {"resp": "<=2h", "ontime": ">=99%", "cust": "Si"},
 1601023310453: {"resp": "<=3h", "ontime": ">=98%", "cust": "Si"},
 1601758712138: {"resp": "<=4h", "ontime": ">=90%", "cust": "Si"},
 1601765195031: {"resp": "<=3h", "ontime": ">=91%", "cust": "Si"},
 1601718581780: {"resp": "<=3h", "ontime": ">=91%", "cust": "Si"},
 1600238919442: {"resp": "<=2h", "ontime": ">=96%", "cust": "Si"},
}

def fnum(s, default=0.0):
    m = re.search(r"[\d.]+", str(s).replace(",", ""))
    return float(m.group()) if m else default

def pid_de(url):
    m = re.search(r"_(\d+)\.html", url)
    return int(m.group(1)) if m else None

rows, seen = [], set()
with open(SRC, encoding="utf-8-sig", newline="") as f:
    for r in csv.DictReader(f):
        u = r.get("url", "").strip()
        if not u or u in seen:
            continue
        seen.add(u)
        ficha = FICHAS.get(pid_de(u), {})
        rows.append({
            "url": u,
            "title": r.get("title", ""),
            "company": r.get("company", ""),
            "country": r.get("country", ""),
            "years": int(fnum(r.get("years", 0))),
            "rating": fnum(r.get("rating", 0)),
            "reviews": int(fnum(r.get("reviews", 0))),
            "service": fnum(r.get("service", 0)),
            "shipping": fnum(r.get("shipping", 0)),
            "price_raw": r.get("price", ""),
            "price_min": fnum(r.get("price", ""), default=1e9),
            "moq": r.get("moq", "").replace("Min. order:", "").strip(),
            "resp": ficha.get("resp", ""),
            "ontime": ficha.get("ontime", ""),
            "cust": ficha.get("cust", ""),
        })

rows.sort(key=lambda x: (-x["reviews"], -x["years"], -x["service"], x["price_min"]))

wb = Workbook(); ws = wb.active; ws.title = "Proveedores squishy"
cols = [("#", 5), ("URL", 60), ("Producto", 32), ("Proveedor", 36), ("Pais", 6),
        ("Reviews", 9), ("Anios", 7), ("Service", 8), ("Shipping", 9),
        ("Rating", 7), ("Precio USD", 12), ("MOQ", 16),
        ("Response", 10), ("On-time", 10), ("Custom/OEM", 11)]
hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True)
fichaFill = PatternFill("solid", fgColor="E8F0FE")
for i, (name, w) in enumerate(cols, 1):
    c = ws.cell(1, i, name); c.fill = hf; c.font = hfont
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = w

for idx, r in enumerate(rows, 1):
    precio = "" if r["price_min"] >= 1e9 else r["price_raw"]
    ws.append([idx, r["url"], r["title"], r["company"], r["country"],
               r["reviews"], r["years"], r["service"], r["shipping"],
               r["rating"], precio, r["moq"], r["resp"], r["ontime"], r["cust"]])
    # resaltar celdas de ficha si hay dato
    if r["resp"]:
        for col in (13, 14, 15):
            ws.cell(idx + 1, col).fill = fichaFill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
wb.save(OUT)
enriquecidas = sum(1 for r in rows if r["resp"])
print(f"OK -> {OUT}  ({len(rows)} filas, {enriquecidas} con datos de ficha)")
